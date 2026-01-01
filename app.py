import io
import re
import difflib
import datetime as dt
from collections import Counter
import numpy as np
import pandas as pd
import openpyxl
import streamlit as st

st.set_page_config(page_title="QC Utilization Summary", layout="wide")
st.title("📊 QC Utilization — Tracker → Summary (and optional Report fill)")
st.caption("Upload the **Daily Tracker** (tabs per agent) → get the **same kind of summary** + optionally a **filled Utilization Report** workbook.")

# ---------------- Helpers ----------------

def normalize_agent(name: str) -> str:
    s = str(name or "").replace("_", " ").replace("-", " ").strip()
    # Insert space between lower->Upper (e.g., AlOmari -> Al Omari)
    s = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def norm_key(s: str) -> str:
    s = normalize_agent(s).lower()
    s = re.sub(r"[^a-z\s]", " ", s)
    s = re.sub(r"\b(al|el)\b", " ", s)   # drop "Al/El" for fuzzy matching
    s = re.sub(r"\s+", " ", s).strip()
    return s

def best_match(name: str, choices: list[str]) -> tuple[str, float]:
    key = norm_key(name)
    best, best_score = "", 0.0
    for c in choices:
        score = difflib.SequenceMatcher(None, key, norm_key(c)).ratio()
        if score > best_score:
            best_score, best = score, c
    return best, best_score

def parse_time_to_hours(x) -> float:
    """Parse Excel-ish time values to hours (float)."""
    if x is None or x == "":
        return 0.0
    if isinstance(x, dt.timedelta):
        return x.total_seconds() / 3600.0
    if isinstance(x, dt.time):
        return (x.hour * 3600 + x.minute * 60 + x.second) / 3600.0
    if isinstance(x, (int, float)):
        v = float(x)
        # Common cases:
        # - Excel time fraction of a day (0..1) => *24
        # - Sometimes already hours (<=24)
        # - Sometimes minutes (>24 and <= 24*60)
        if v <= 1.0:
            return v * 24.0
        if v <= 24.0:
            return v
        if v <= 24.0 * 60.0:
            return v / 60.0
        return v / 3600.0
    if isinstance(x, str):
        s = x.strip()
        m = re.match(r"^(\d+)\s*:\s*(\d{1,2})(?:\s*:\s*(\d{1,2}))?$", s)
        if m:
            h = int(m.group(1)); mi = int(m.group(2)); sec = int(m.group(3) or 0)
            return (h * 3600 + mi * 60 + sec) / 3600.0
        m = re.match(r"^(\d+(?:\.\d+)?)\s*h$", s.lower())
        if m:
            return float(m.group(1))
        m = re.match(r"^(\d+)\s*m$", s.lower())
        if m:
            return int(m.group(1)) / 60.0
    return 0.0

def categorize(ticket_type, ticket_id) -> str | None:
    """
    Map tracker rows -> report categories.
    Categories (match your report logic):
      shops, with_links, no_links, ext_pos, new_outlet,
      large_update (replace menu), medium_update, small_update,
      whatsapp, special_project, unknown
    """
    tt = "" if ticket_type is None else str(ticket_type).strip()
    tid = "" if ticket_id is None else str(ticket_id).strip()

    ttn = re.sub(r"\s+", " ", tt).strip().lower()
    tidn = tid.lower()

    # Skip obvious junk that sometimes lands in Ticket Type column
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", ttn) or re.fullmatch(r"\d{1,2}:\d{2}.*", ttn):
        return None
    if ttn in ("", "none", "nan"):
        return None

    # WhatsApp can show up in either Ticket Type or Ticket#
    if ("whatsapp" in ttn) or ("whats app" in ttn) or ("whatsapp" in tidn) or ("whats" in tidn):
        return "whatsapp"

    # Special tasks/projects/scores/training
    if any(k in ttn for k in ["special", "project", "task", "tasks", "score", "scores", "meeting", "training", "weekly", "monthly", "tp scores", "utilization"]):
        return "special_project"

    if "new outlet" in ttn:
        return "new_outlet"

    # Builds
    if ("shop" in ttn) or ("brand setup" in ttn) or ("new brand" in ttn) or ("existing brand" in ttn):
        return "shops"
    if ("with links" in ttn) or ("scrap" in ttn) or ("scrapping" in ttn):
        return "with_links"
    if ("no links" in ttn) or ("non pos" in ttn) or ("non-pos" in ttn):
        return "no_links"
    if ("pos" in ttn) or ("ext/pos" in ttn) or ("existing pos" in ttn) or ("pos build" in ttn):
        return "ext_pos"

    # Updates
    if ("replace" in ttn) or ("menu replace" in ttn) or ("menu replacement" in ttn):
        return "large_update"
    if "large update" in ttn:
        return "large_update"
    if "medium" in ttn:
        return "medium_update"
    if "small" in ttn:
        return "small_update"

    # Fallback: generic update words
    if any(k in ttn for k in ["menu update","update","price","add on","add ons","new items","remove items","disable","image update","change status","update info","operational hours"]):
        return "medium_update"

    return "unknown"

def hours_to_timedelta(h: float) -> dt.timedelta:
    return dt.timedelta(seconds=int(round(float(h) * 3600)))

def extract_targets_from_report(wb: openpyxl.Workbook) -> dict:
    """
    Pull target average times (hours) from Summary row 2.
    Falls back to the defaults found in your uploaded template.
    """
    defaults = {
        "shops": 4.0,
        "with_links": 2.5,
        "no_links": 3.0,
        "ext_pos": 0.5,
        "medium_update": 1.0,
        "large_update": 2.0,
        "small_update": 0.5,
        "whatsapp": 0.13333333333333333,  # 8 minutes
    }
    if "Summary" not in wb.sheetnames:
        return defaults

    ws = wb["Summary"]
    headers = [ws.cell(1, c).value for c in range(1, 250)]
    # trim
    while headers and headers[-1] is None:
        headers.pop()

    def col_idx(name: str) -> int | None:
        for i, h in enumerate(headers, start=1):
            if str(h).strip() == name:
                return i
        return None

    mapping = {
        "shops": "Avg Shops",
        "with_links": "Avg With links",
        "no_links": "Avg No links",
        "ext_pos": "Avg for Ext/POS",
        "medium_update": "Avg for medium update",
        "large_update": "Avg for Large update",
        "small_update": "Avg for Small update",
        "whatsapp": "Avg for Whatsapp",
    }

    out = {}
    for k, header in mapping.items():
        ci = col_idx(header)
        if not ci:
            out[k] = defaults[k]
            continue
        val = ws.cell(2, ci).value
        out[k] = parse_time_to_hours(val) if val is not None else defaults[k]

    # sanity fallbacks
    for k, v in defaults.items():
        if k not in out or not isinstance(out[k], (int, float)) or out[k] <= 0:
            out[k] = v
    return out

@st.cache_data(show_spinner=False)
def parse_tracker(tracker_bytes: bytes) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      - rows_df: row-level records
      - unknown_df: unknown ticket types frequency table
    """
    wb = openpyxl.load_workbook(io.BytesIO(tracker_bytes), data_only=True)
    records = []
    unknown_counter = Counter()

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        agent = normalize_agent(sheet)

        # Expect columns: A=Ticket Type, B=Ticket#, C=Time Taken, D=Date
        for r in range(2, ws.max_row + 1):
            date_cell = ws.cell(r, 4).value
            if not isinstance(date_cell, (dt.datetime, dt.date)):
                continue
            date = date_cell.date() if isinstance(date_cell, dt.datetime) else date_cell

            ticket_type = ws.cell(r, 1).value
            ticket_id = ws.cell(r, 2).value
            time_taken = ws.cell(r, 3).value

            cat = categorize(ticket_type, ticket_id)
            if cat is None:
                continue

            hours = parse_time_to_hours(time_taken)

            if cat == "unknown":
                unknown_counter[str(ticket_type).strip()] += 1

            records.append({
                "agent_sheet": sheet,
                "agent": agent,
                "date": date,
                "ticket_type": ticket_type,
                "ticket_id": ticket_id,
                "hours": hours,
                "category": cat,
            })

    rows_df = pd.DataFrame(records)
    unknown_df = pd.DataFrame(
        [{"ticket_type": k, "count": v} for k, v in unknown_counter.most_common()]
    )
    return rows_df, unknown_df

def build_daily_agg(rows_df: pd.DataFrame) -> pd.DataFrame:
    g = rows_df.groupby(["agent", "date"])
    out = g.apply(lambda x: pd.Series({
        "hours_worked": x["hours"].sum(),
        "shops": int((x["category"] == "shops").sum()),
        "with_links": int((x["category"] == "with_links").sum()),
        "no_links": int((x["category"] == "no_links").sum()),
        "ext_pos": int((x["category"] == "ext_pos").sum()),
        "new_outlet": int((x["category"] == "new_outlet").sum()),
        "large_update": int((x["category"] == "large_update").sum()),
        "medium_update": int((x["category"] == "medium_update").sum()),
        "small_update": int((x["category"] == "small_update").sum()),
        "whatsapp": int((x["category"] == "whatsapp").sum()),
        "special_project_hours": x.loc[x["category"] == "special_project", "hours"].sum(),
    })).reset_index()
    return out

def build_summary(daily_agg: pd.DataFrame, targets: dict, avail_hours_per_day: float, group_by: str) -> pd.DataFrame:
    """
    group_by: "month" or "week"
    """
    df = daily_agg.copy()
    df["available_hours"] = avail_hours_per_day

    if group_by == "week":
        # ISO week (Mon-Sun). More consistent than Excel's locale weeknum differences.
        df["year"] = pd.to_datetime(df["date"]).dt.isocalendar().year.astype(int)
        df["week"] = pd.to_datetime(df["date"]).dt.isocalendar().week.astype(int)
        key_cols = ["agent", "year", "week"]
    else:
        df["year"] = pd.to_datetime(df["date"]).dt.year.astype(int)
        df["month"] = pd.to_datetime(df["date"]).dt.month.astype(int)
        key_cols = ["agent", "year", "month"]

    agg = df.groupby(key_cols).agg(
        workdays=("date", "nunique"),
        available_hours=("available_hours", lambda s: float(s.iloc[0]) * s.size),
        hours_worked=("hours_worked", "sum"),
        shops=("shops", "sum"),
        with_links=("with_links", "sum"),
        no_links=("no_links", "sum"),
        ext_pos=("ext_pos", "sum"),
        new_outlet=("new_outlet", "sum"),
        large_update=("large_update", "sum"),
        medium_update=("medium_update", "sum"),
        small_update=("small_update", "sum"),
        whatsapp=("whatsapp", "sum"),
        special_project_hours=("special_project_hours", "sum"),
    ).reset_index()

    agg["builds"] = agg["shops"] + agg["with_links"] + agg["no_links"]
    agg["updates"] = agg["medium_update"] + agg["small_update"] + agg["large_update"]
    agg["tickets"] = agg["builds"] + agg["updates"] + agg["ext_pos"] + agg["new_outlet"] + agg["whatsapp"]
    agg["utilization"] = np.where(agg["available_hours"] > 0, agg["hours_worked"] / agg["available_hours"], np.nan)

    agg["target_duration_hours"] = (
        agg["shops"] * targets["shops"]
        + agg["with_links"] * targets["with_links"]
        + agg["no_links"] * targets["no_links"]
        + (agg["ext_pos"] + agg["new_outlet"]) * targets["ext_pos"]
        + agg["medium_update"] * targets["medium_update"]
        + agg["large_update"] * targets["large_update"]
        + agg["small_update"] * targets["small_update"]
        + agg["whatsapp"] * targets["whatsapp"]
        + agg["special_project_hours"]
    )
    agg["result"] = np.where(agg["target_duration_hours"] >= agg["hours_worked"], True, "Check")

    # nicer ordering
    front = key_cols + ["workdays", "available_hours", "hours_worked", "utilization", "tickets", "builds", "updates", "special_project_hours",
                        "shops", "with_links", "no_links", "ext_pos", "new_outlet", "medium_update", "large_update", "small_update", "whatsapp",
                        "target_duration_hours", "result"]
    agg = agg[front].sort_values(["utilization", "hours_worked"], ascending=[False, False])
    return agg

def fill_report_workbook(report_bytes: bytes, month_sheet: str, daily_agg: pd.DataFrame, tracker_agents: list[str]) -> tuple[bytes, pd.DataFrame]:
    wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
    if month_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{month_sheet}' not found in report workbook.")

    ws = wb[month_sheet]

    # Pull employee names from report sheet
    report_names = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        d = ws.cell(r, 2).value
        if isinstance(v, str) and isinstance(d, dt.datetime):
            report_names.append(v.strip())
    report_names = sorted(set(report_names))

    # Auto-match report name -> tracker agent name
    matches = []
    for rn in report_names:
        best, score = best_match(rn, tracker_agents)
        matches.append({"report_name": rn, "tracker_agent": best, "match_score": round(score, 3)})
    match_df = pd.DataFrame(matches).sort_values("match_score")

    # Build lookup (agent, date) -> daily metrics
    lookup = {}
    for row in daily_agg.itertuples(index=False):
        lookup[(row.agent, row.date)] = row

    filled = 0
    missing = 0

    # Column map in your report month sheet (based on your template)
    # A=Employee, B=Date, F=Hours Worked, G=Build Shops, H=Build Scrapping, I=Build Non POS,
    # K=Build POS/outlet, L=New outlet/outlet, M=Replace menu, N=Update medium, O=Update small,
    # P=Whatsapp, Q=Special project
    for r in range(2, ws.max_row + 1):
        emp = ws.cell(r, 1).value
        d = ws.cell(r, 2).value
        if not isinstance(emp, str) or not isinstance(d, dt.datetime):
            continue
        emp = emp.strip()
        date = d.date()

        # resolve emp -> tracker agent
        rec = match_df.loc[match_df["report_name"] == emp]
        if rec.empty:
            missing += 1
            continue
        agent = rec.iloc[0]["tracker_agent"]

        data = lookup.get((agent, date))
        if not data:
            missing += 1
            continue

        ws.cell(r, 6).value = hours_to_timedelta(data.hours_worked) if data.hours_worked else None
        ws.cell(r, 7).value = float(data.shops) if data.shops else None
        ws.cell(r, 8).value = float(data.with_links) if data.with_links else None
        ws.cell(r, 9).value = float(data.no_links) if data.no_links else None
        ws.cell(r, 11).value = float(data.ext_pos) if data.ext_pos else None
        ws.cell(r, 12).value = float(data.new_outlet) if data.new_outlet else None
        ws.cell(r, 13).value = float(data.large_update) if data.large_update else None
        ws.cell(r, 14).value = float(data.medium_update) if data.medium_update else None
        ws.cell(r, 15).value = float(data.small_update) if data.small_update else None
        ws.cell(r, 16).value = float(data.whatsapp) if data.whatsapp else None
        ws.cell(r, 17).value = hours_to_timedelta(data.special_project_hours) if data.special_project_hours else None

        filled += 1

    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue(), match_df

# ---------------- UI ----------------

col1, col2, col3 = st.columns([1.2, 1.2, 1.2])

with col1:
    tracker_file = st.file_uploader("1) Upload **QC Team Daily Tracker** (xlsx)", type=["xlsx"])
with col2:
    report_file = st.file_uploader("2) (Optional) Upload **QC Utilization Report** (xlsx)", type=["xlsx"])
with col3:
    avail_hours_per_day = st.number_input("Default available hours per day", min_value=1.0, max_value=16.0, value=8.0, step=0.5)

if not tracker_file:
    st.info("Upload the Daily Tracker first 👆")
    st.stop()

tracker_bytes = tracker_file.getvalue()
rows_df, unknown_df = parse_tracker(tracker_bytes)

if rows_df.empty:
    st.error("I couldn’t find any valid rows in the tracker. Make sure the agent tabs have columns: Ticket Type | Ticket# | Time Taken | Date")
    st.stop()

# date filter
min_date = rows_df["date"].min()
max_date = rows_df["date"].max()
date_from, date_to = st.date_input("Date range to summarize", value=(min_date, max_date), min_value=min_date, max_value=max_date)

if isinstance(date_from, tuple) or isinstance(date_to, tuple):
    # streamlit sometimes returns tuple for range; normalize
    date_from, date_to = date_from[0], date_from[1]

rows_df = rows_df[(rows_df["date"] >= date_from) & (rows_df["date"] <= date_to)]
daily_agg = build_daily_agg(rows_df)

# Targets
targets = None
if report_file:
    try:
        report_bytes = report_file.getvalue()
        wb_tmp = openpyxl.load_workbook(io.BytesIO(report_bytes), data_only=True)
        targets = extract_targets_from_report(wb_tmp)
    except Exception:
        targets = None
if targets is None:
    targets = {
        "shops": 4.0,
        "with_links": 2.5,
        "no_links": 3.0,
        "ext_pos": 0.5,
        "medium_update": 1.0,
        "large_update": 2.0,
        "small_update": 0.5,
        "whatsapp": 0.13333333333333333,
    }

with st.expander("🎯 Target average times used in the calculation", expanded=False):
    st.write(pd.DataFrame([targets]).rename(index={0: "hours"}))

group_by = st.radio("Summary level", options=["Month (agent-level)", "Week (agent-level)"], horizontal=True)
group_key = "month" if group_by.startswith("Month") else "week"

summary_df = build_summary(daily_agg, targets, avail_hours_per_day, group_key)

# Quick team insights
if not summary_df.empty:
    best_row = summary_df.iloc[0]
    worst_row = summary_df.sort_values("utilization", ascending=True).iloc[0]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Team hours worked", f"{summary_df['hours_worked'].sum():.1f}h")
    c2.metric("Avg utilization", f"{summary_df['utilization'].mean():.2%}")
    c3.metric("Highest utilization", f"{best_row['agent']} ({best_row['utilization']:.2%})")
    c4.metric("Lowest utilization", f"{worst_row['agent']} ({worst_row['utilization']:.2%})")

st.subheader("✅ Summary")
st.dataframe(summary_df, use_container_width=True)

# Downloads: summary
csv_buf = io.StringIO()
summary_df.to_csv(csv_buf, index=False)
st.download_button("⬇️ Download summary CSV", data=csv_buf.getvalue().encode("utf-8"), file_name="qc_utilization_summary.csv", mime="text/csv")

# Unknown ticket types
if unknown_df is not None and not unknown_df.empty:
    st.subheader("⚠️ Unknown ticket types (not mapped yet)")
    st.caption("These are ticket types that didn’t match any category. If these matter, tell me and I’ll add them to the mapping rules.")
    st.dataframe(unknown_df, use_container_width=True)
    u_csv = io.StringIO()
    unknown_df.to_csv(u_csv, index=False)
    st.download_button("⬇️ Download unknown types CSV", data=u_csv.getvalue().encode("utf-8"), file_name="unknown_ticket_types.csv", mime="text/csv")

# Optional: fill report workbook
if report_file:
    st.subheader("📄 Fill your Utilization Report (optional)")
    try:
        wb_r = openpyxl.load_workbook(io.BytesIO(report_file.getvalue()), data_only=False)
        month_sheets = [s for s in wb_r.sheetnames if s.lower() != "summary"]
        month_sheet = st.selectbox("Pick the month sheet to fill", options=month_sheets, index=0)
        tracker_agents = sorted(rows_df["agent"].unique().tolist())

        if st.button("Generate filled report workbook", type="primary"):
            filled_bytes, match_df = fill_report_workbook(report_file.getvalue(), month_sheet, daily_agg, tracker_agents)
            st.success("Done ✅ Download below.")
            st.download_button("⬇️ Download filled report (xlsx)", data=filled_bytes, file_name=f"QC_Utilization_Report_Filled_{month_sheet}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with st.expander("Name matching used (Report → Tracker)", expanded=False):
                st.dataframe(match_df, use_container_width=True)

    except Exception as e:
        st.error(f"Could not fill report workbook: {e}")
